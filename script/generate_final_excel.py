#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
最终Excel生成器
合并所有批次JSON文件并生成符合规范的Excel测试用例文件
"""

import json
import pandas as pd
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import os
import platform
import traceback

class FinalExcelGenerator:
    def __init__(self):
        self.output_dir = Path("output")
        self.batch_files = []
        self.all_test_cases = []
        self.supported_formats = ['module_*.json', 'batch_*.json']
        
    def detect_document_name(self):
        """检测产品文档名称，用于生成Excel文件名"""
        # 首先尝试从markdown目录获取文档名
        markdown_dir = Path("product/markdown")
        if markdown_dir.exists():
            md_files = list(markdown_dir.glob("*.md"))
            if md_files:
                # 取第一个md文件名作为产品名
                md_file = md_files[0]
                product_name = md_file.stem.replace('_', '')
                return product_name
        
        # 其次尝试从product目录获取docx文件名
        product_dir = Path("product")
        if product_dir.exists():
            docx_files = list(product_dir.glob("*.docx"))
            if docx_files:
                docx_file = docx_files[0]
                product_name = docx_file.stem.replace('_', '')
                return product_name
        
        # 默认名称
        return "测试用例"

    def scan_batch_files(self):
        """扫描所有批次JSON文件，统一使用module格式"""
        print("🔍 扫描测试用例文件...")
        
        # 优先扫描新的模块文件格式 module_xx_模块名.json
        module_files = list(self.output_dir.glob("module_*.json"))
        
        # 如果没有模块文件，检查是否有传统批次文件需要转换
        if not module_files:
            print("📄 检查传统批次文件...")
            batch_files = list(self.output_dir.glob("batch_*.json"))
            if batch_files:
                print("⚠️  发现传统批次文件，建议转换为模块格式")
                print("   使用模块格式可以更好地组织测试用例")
                self.batch_files = sorted(batch_files)
            else:
                print("❌ 未发现任何测试用例文件")
                return
        else:
            print("📁 使用模块文件格式")
            # 按模块编号排序
            self.batch_files = sorted(module_files, key=lambda x: self._extract_module_number(x.name))
            
        print(f"📊 发现 {len(self.batch_files)} 个文件:")
        for file in self.batch_files:
            print(f"  - {file.name}")
    
    def _extract_module_number(self, filename):
        """从文件名中提取模块编号用于排序"""
        import re
        match = re.search(r'module_(\d+)', filename)
        return int(match.group(1)) if match else 999
            
    def load_and_merge_data(self):
        """加载并合并所有批次数据"""
        print("\n📥 加载并合并数据...")
        
        total_cases = 0
        modules_processed = set()
        
        for batch_file in self.batch_files:
            try:
                with open(batch_file, 'r', encoding='utf-8') as f:
                    batch_data = json.load(f)
                    
                # 支持两种格式：数组格式和对象格式
                if isinstance(batch_data, list):
                    test_cases = batch_data
                else:
                    batch_info = batch_data.get('batch_info', {})
                    test_cases = batch_data.get('test_cases', [])
                
                print(f"  ✅ {batch_file.name}: {len(test_cases)} 个用例")
                
                for case in test_cases:
                    # 确保所有必需字段存在
                    # 用例等级处理：只显示P1，P2和P3显示为空白
                    original_level = case.get('用例等级', '')
                    display_level = original_level if original_level == 'P1' else ''
                    
                    case_data = {
                        '测试模块': case.get('测试模块', ''),
                        '用例名称': case.get('用例名称', ''),
                        '前置条件': case.get('前置条件', ''),
                        '用例步骤': case.get('用例步骤', ''),
                        '预期结果': case.get('预期结果', ''),
                        '用例等级': display_level,  # 只显示P1，其他为空
                        '开发验证人员': case.get('开发验证人员', ''),
                        '冒烟测试结果': case.get('自测验收结果', ''),
                        '测试人员': case.get('测试人员', ''),
                        '测试结果-安卓': case.get('测试结果-安卓', ''),
                        '测试结果-iOS': case.get('测试结果-iOS', ''),
                        '需求原文': case.get('需求原文', ''),  # 新增需求原文字段
                        '备注': case.get('备注', '')
                    }
                    
                    # 保存原始等级用于统计
                    case_data['_original_level'] = original_level
                    
                    self.all_test_cases.append(case_data)
                    modules_processed.add(case_data['测试模块'])
                    
                total_cases += len(test_cases)
                
            except json.JSONDecodeError as e:
                print(f"  ❌ {batch_file.name}: JSON格式错误 - {e}")
                print(f"     错误位置: 行{e.lineno}, 列{e.colno}")
            except FileNotFoundError as e:
                print(f"  ❌ {batch_file.name}: 文件不存在 - {e}")
            except PermissionError as e:
                print(f"  ❌ {batch_file.name}: 权限不足 - {e}")
            except Exception as e:
                print(f"  ❌ {batch_file.name}: 加载失败 - {e}")
                print(f"     错误类型: {type(e).__name__}")
                
        print(f"\n📊 合并完成:")
        print(f"  - 总用例数: {total_cases}")
        print(f"  - 模块数: {len(modules_processed)}")
        print(f"  - 模块列表: {', '.join(sorted(modules_processed))}")
        
    def create_excel_file(self):
        """创建Excel文件"""
        print("\n📊 生成Excel文件...")
        
        if not self.all_test_cases:
            print("❌ 没有测试用例数据")
            return None
        
        # 检测产品名称
        product_name = self.detect_document_name()
        print(f"📋 产品名称: {product_name}")
            
        # 创建DataFrame
        df = pd.DataFrame(self.all_test_cases)
        
        # 创建一个用于统计的完整DataFrame（包含_original_level）
        stats_df = df.copy()
        
        # 从显示的DataFrame中删除临时列
        display_df = df.drop('_original_level', axis=1)
        
        # 生成动态文件名
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_filename = f"{product_name}_测试用例_{timestamp}.xlsx"
        excel_path = self.output_dir / excel_filename
        
        # 创建Excel文件
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            display_df.to_excel(writer, sheet_name='测试用例', index=False)
            
            # 获取工作表
            worksheet = writer.sheets['测试用例']
            
            # 应用格式化（使用显示的DataFrame）
            self._apply_formatting(worksheet, display_df)
            
            # 添加统计信息工作表（使用包含原始等级的DataFrame）
            self._create_statistics_sheet(writer, stats_df)
            
        print(f"✅ Excel文件已生成: {excel_path}")
        return excel_path
        
    def _apply_formatting(self, worksheet, df):
        """应用Excel格式化"""
        print("🎨 应用Excel格式化...")
        
        # 定义样式
        header_font = Font(name='微软雅黑', size=12, bold=True, color='000000')
        header_fill = PatternFill(start_color='B3D600', end_color='B3D600', fill_type='solid')  # 黄色背景
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        cell_font = Font(name='微软雅黑', size=10, color='000000')
        cell_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        thin_border = Border(
            left=Side(style='thin', color='C0C0C0'),
            right=Side(style='thin', color='C0C0C0'),
            top=Side(style='thin', color='C0C0C0'),
            bottom=Side(style='thin', color='C0C0C0')
        )
        
        # 自适应列宽设置 - 根据内容自动调整
        self._auto_adjust_column_width(worksheet, df)
        
        # 设置标题行高度
        worksheet.row_dimensions[1].height = 25
            
        # 格式化标题行
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
        # 格式化数据行
        for row in range(2, len(df) + 2):
            # 自适应行高
            self._auto_adjust_row_height(worksheet, row, df)
            
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.font = cell_font
                cell.alignment = cell_alignment
                cell.border = thin_border
                
                # 用例等级特殊颜色
                if col == 6:  # 用例等级列
                    if cell.value == 'P1':
                        cell.font = Font(name='微软雅黑', size=10, color='FF0000', bold=True)
                
                # 测试结果列的条件格式
                if col in [10, 11]:  # 测试结果-安卓和iOS列
                    self._apply_test_result_formatting(cell)
                        
        # 冻结首行
        worksheet.freeze_panes = 'A2'
        
        # 添加下拉列表验证
        self._add_dropdown_validation(worksheet, df)
        
        # 合并相同前置条件的单元格
        self._merge_same_preconditions(worksheet, df)
        
        # 合并相同测试模块的单元格
        self._merge_same_test_modules(worksheet, df)
    
    def _auto_adjust_column_width(self, worksheet, df):
        """自适应列宽设置"""
        print("📏 自适应调整列宽...")
        
        # 列标题
        column_headers = list(df.columns)
        
        for col_idx, header in enumerate(column_headers, start=1):
            # 获取列字母
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            
            # 计算标题长度
            max_length = len(str(header))
            
            # 遍历该列所有数据，找到最长的内容
            for row_idx in range(2, len(df) + 2):
                cell_value = worksheet.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    # 计算内容长度，考虑换行符
                    lines = str(cell_value).split('\n')
                    max_line_length = max(len(line) for line in lines) if lines else 0
                    max_length = max(max_length, max_line_length)
            
            # 设置列宽，考虑中文字符和最大最小值限制
            # 中文字符按1.5倍计算，英文字符按1倍计算
            adjusted_length = max_length * 1.2  # 稍微增加一些边距
            
            # 设置最小和最大列宽限制
            min_width = 8
            max_width = 50
            
            # 针对不同列设置不同的最大宽度
            if header in ['用例步骤', '预期结果']:
                max_width = 60
            elif header in ['需求原文']:
                max_width = 80  # 需求原文列较宽，包含定位信息和原文内容
            elif header in ['前置条件', '备注']:
                max_width = 40
            elif header in ['用例名称']:
                max_width = 35
            elif header in ['用例等级', '开发验证人员', '测试人员', '冒烟测试结果', '测试结果-安卓', '测试结果-iOS']:
                max_width = 15
            
            final_width = max(min_width, min(adjusted_length, max_width))
            worksheet.column_dimensions[col_letter].width = final_width
    
    def _auto_adjust_row_height(self, worksheet, row_idx, df):
        """自适应行高设置"""
        max_lines = 1
        
        # 遍历该行所有单元格，找到最多的行数
        for col_idx in range(1, len(df.columns) + 1):
            cell_value = worksheet.cell(row=row_idx, column=col_idx).value
            if cell_value:
                lines = str(cell_value).count('\n') + 1
                max_lines = max(max_lines, lines)
        
        # 根据行数设置行高，每行约20像素
        base_height = 20
        row_height = max(base_height, max_lines * 18 + 6)  # 6像素的上下边距
        
        # 设置最大行高限制
        max_height = 100
        final_height = min(row_height, max_height)
        
        worksheet.row_dimensions[row_idx].height = final_height
    
    def _apply_test_result_formatting(self, cell):
        """为测试结果单元格应用条件格式"""
        if cell.value:
            if str(cell.value).upper() == 'PASS':
                cell.fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # 绿色
                cell.font = Font(name='微软雅黑', size=10, color='000000', bold=True)
            elif str(cell.value).upper() == 'FAILED':
                cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')  # 红色
                cell.font = Font(name='微软雅黑', size=10, color='FFFFFF', bold=True)
    
    def _add_dropdown_validation(self, worksheet, df):
        """为测试结果列添加下拉列表验证"""
        print("📋 添加下拉列表验证...")
        
        # 创建数据验证对象
        test_result_validation = DataValidation(
            type="list",
            formula1='"PASS,FAILED"',
            allow_blank=True
        )
        test_result_validation.error = '请选择PASS或FAILED'
        test_result_validation.errorTitle = '输入错误'
        test_result_validation.prompt = '请从下拉列表中选择测试结果'
        test_result_validation.promptTitle = '测试结果'
        
        # 添加验证到工作表
        worksheet.add_data_validation(test_result_validation)
        
        # 应用到测试结果-安卓列（第10列，J列）
        android_range = f"J2:J{len(df) + 1}"
        test_result_validation.add(android_range)
        
        # 应用到测试结果-iOS列（第11列，K列）
        ios_range = f"K2:K{len(df) + 1}"
        test_result_validation.add(ios_range)
        
    def _merge_same_preconditions(self, worksheet, df):
        """合并相同前置条件的单元格"""
        print("🔗 合并相同前置条件...")
        
        precondition_col = 3  # 前置条件列
        current_condition = None
        start_row = 2
        
        for row in range(2, len(df) + 2):
            cell_value = worksheet.cell(row=row, column=precondition_col).value
            
            if cell_value != current_condition:
                # 如果前置条件改变，合并之前的单元格
                if current_condition is not None and row - start_row > 1:
                    try:
                        worksheet.merge_cells(f'C{start_row}:C{row-1}')
                        # 设置合并单元格的对齐方式
                        merged_cell = worksheet.cell(row=start_row, column=precondition_col)
                        merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    except:
                        pass  # 忽略合并错误
                        
                current_condition = cell_value
                start_row = row
                
        # 处理最后一组
        if current_condition is not None and len(df) + 1 - start_row >= 1:
            try:
                if len(df) + 1 - start_row > 0:
                    worksheet.merge_cells(f'C{start_row}:C{len(df) + 1}')
                    merged_cell = worksheet.cell(row=start_row, column=precondition_col)
                    merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            except:
                pass  # 忽略合并错误
    
    def _merge_same_test_modules(self, worksheet, df):
        """合并相同测试模块的单元格"""
        print("🔗 合并相同测试模块...")
        
        test_module_col = 1  # 测试模块列（第一列）
        current_module = None
        start_row = 2
        
        for row in range(2, len(df) + 2):
            cell_value = worksheet.cell(row=row, column=test_module_col).value
            
            if cell_value != current_module:
                # 如果测试模块改变，合并之前的单元格
                if current_module is not None and row - start_row > 1:
                    try:
                        worksheet.merge_cells(f'A{start_row}:A{row-1}')
                        # 设置合并单元格的对齐方式
                        merged_cell = worksheet.cell(row=start_row, column=test_module_col)
                        merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    except:
                        pass  # 忽略合并错误
                        
                current_module = cell_value
                start_row = row
                
        # 处理最后一组
        if current_module is not None and len(df) + 1 - start_row >= 1:
            try:
                if len(df) + 1 - start_row > 0:
                    worksheet.merge_cells(f'A{start_row}:A{len(df) + 1}')
                    merged_cell = worksheet.cell(row=start_row, column=test_module_col)
                    merged_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            except:
                pass  # 忽略合并错误
    
    def _create_statistics_sheet(self, writer, df):
        """创建统计信息工作表"""
        print("📈 生成统计信息...")
        
        stats_data = []
        
        # 总体统计
        total_cases = len(df)
        stats_data.append(['总测试用例数', total_cases])
        stats_data.append(['', ''])  # 空行
        
        # 按模块统计
        module_stats = df['测试模块'].value_counts()
        stats_data.append(['模块统计', '用例数'])
        for module, count in module_stats.items():
            stats_data.append([module, count])
        stats_data.append(['', ''])  # 空行
        
        # 按优先级统计（使用原始等级数据）
        priority_stats = df['_original_level'].value_counts()
        stats_data.append(['优先级统计', '用例数'])
        for priority in ['P1', 'P2', 'P3']:
            count = priority_stats.get(priority, 0)
            stats_data.append([priority, count])
        stats_data.append(['', ''])  # 空行
        
        # 生成时间
        stats_data.append(['生成时间', datetime.now().strftime("%Y-%m-%d %H:%M:%S")])
        
        # 创建统计DataFrame
        stats_df = pd.DataFrame(stats_data, columns=['项目', '数值'])
        stats_df.to_excel(writer, sheet_name='统计信息', index=False)
        
        # 格式化统计表
        stats_worksheet = writer.sheets['统计信息']
        
        # 设置列宽
        stats_worksheet.column_dimensions['A'].width = 20
        stats_worksheet.column_dimensions['B'].width = 15
        
        # 格式化标题行
        header_font = Font(name='微软雅黑', size=12, bold=True)
        for col in range(1, 3):
            cell = stats_worksheet.cell(row=1, column=col)
            cell.font = header_font
                
    def generate_statistics(self):
        """生成统计信息"""
        if not self.all_test_cases:
            return {}
            
        stats = {
            'total_cases': len(self.all_test_cases),
            'modules': {},
            'priorities': {'P1': 0, 'P2': 0, 'P3': 0}
        }
        
        for case in self.all_test_cases:
            module = case['测试模块']
            priority = case.get('_original_level', case['用例等级'])  # 使用原始等级进行统计
            
            if module not in stats['modules']:
                stats['modules'][module] = 0
            stats['modules'][module] += 1
            
            if priority in stats['priorities']:
                stats['priorities'][priority] += 1
                
        return stats
        
    def run(self):
        """执行完整流程"""
        print("🚀 开始生成最终Excel文件...")
        
        try:
            # 1. 扫描批次文件
            self.scan_batch_files()
            
            if not self.batch_files:
                print("❌ 未发现任何批次文件")
                return None
                
            # 2. 加载并合并数据
            self.load_and_merge_data()
            
            if not self.all_test_cases:
                print("❌ 没有有效的测试用例数据")
                return None
                
            # 3. 生成统计信息
            stats = self.generate_statistics()
            print(f"\n📈 统计信息:")
            print(f"  - 总用例数: {stats['total_cases']}")
            print(f"  - P1用例: {stats['priorities']['P1']}")
            print(f"  - P2用例: {stats['priorities']['P2']}")
            print(f"  - P3用例: {stats['priorities']['P3']}")
            
            # 4. 创建Excel文件
            excel_path = self.create_excel_file()
            
            if excel_path:
                print(f"\n✅ 处理完成!")
                print(f"📊 Excel文件: {excel_path}")
                return excel_path
            else:
                print("❌ Excel生成失败")
                return None
                
        except PermissionError as e:
            print(f"❌ 权限错误: {e}")
            print("请检查output目录的写入权限，或关闭已打开的Excel文件")
            return None
        except FileNotFoundError as e:
            print(f"❌ 文件不存在: {e}")
            print("请确认output目录和测试用例文件存在")
            return None
        except Exception as e:
            print(f"❌ 处理失败: {e}")
            print(f"错误类型: {type(e).__name__}")
            print("详细错误信息:")
            traceback.print_exc()
            return None

def main():
    generator = FinalExcelGenerator()
    return generator.run()

if __name__ == "__main__":
    main() 